import os
import sys
import datetime
from pathlib import Path
from flask import (Flask, render_template, request, flash,
                   redirect, url_for, send_file, abort)
import openpyxl
from openpyxl.styles import Font as OpenpyxlFont
from fpdf import FPDF

# --- Configuração da Aplicação Flask ---
app = Flask(__name__)
# É crucial definir uma chave secreta para usar 'flash'
app.config['SECRET_KEY'] = os.urandom(24) # Gera uma chave aleatória a cada execução

# --- Configuração de Caminhos (Adaptado para Flask) ---
if getattr(sys, 'frozen', False):
    # Se executando como bundle (.exe)
    application_path = Path(sys.executable).parent
else:
    # Se executando como script .py
    application_path = Path(__file__).parent

EXCEL_FILE_PATH = application_path / "Relatorio_Ocorrencias_Python.xlsx"
PDF_SAVE_FOLDER = application_path / "PDFs_Ocorrencias_Python"
PDF_SAVE_FOLDER.mkdir(exist_ok=True) # Cria a pasta se não existir

# Lista de opções para o campo 'Tipo' (para passar ao template)
TIPOS_OCORRENCIA = ["Não Conformidade", "Oportunidade de Melhoria", "Incidente", "Observação", "Outro"]

# --- Funções Lógicas (Reutilizadas e Adaptadas) ---

def _get_excel_headers(data_dict):
    """Define a ordem e os nomes exatos das colunas para o Excel."""
    # Incluir Timestamp primeiro
    headers = ['Timestamp'] + list(data_dict.keys())
    # Remover duplicados caso 'Timestamp' já esteja em data_dict (não deve acontecer)
    headers = sorted(set(headers), key=lambda x: headers.index(x))
    # Remover o campo 'action' se existir (veio do botão)
    if 'action' in headers:
        headers.remove('action')
    return headers


def _save_data_logic(data_dict):
    """Lógica para salvar os dados no Excel. Retorna True/False."""
    try:
        workbook = None
        sheet = None
        headers = _get_excel_headers(data_dict) # Usa a função para obter cabeçalhos

        if EXCEL_FILE_PATH.exists():
            try:
                workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
                sheet_name = "Ocorrencias"
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook.create_sheet(sheet_name)
                    sheet.append(headers) # Cabeçalhos se folha nova
                    for col_num, header in enumerate(headers, 1):
                         sheet.cell(row=1, column=col_num).font = OpenpyxlFont(bold=True)

            except Exception as load_error:
                 print(f"Erro ao carregar Excel existente: {load_error}")
                 flash(f"Erro ao carregar ficheiro Excel: {load_error}", "error")
                 return False # Falha no carregamento
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Ocorrencias"
            sheet.append(headers) # Cabeçalhos
            for col_num, header in enumerate(headers, 1):
                 sheet.cell(row=1, column=col_num).font = OpenpyxlFont(bold=True)


        # Adicionar dados
        next_row = sheet.max_row + 1
        # Garantir que todos os headers existem no data_dict, mesmo que vazios
        # E na ordem correta dos headers
        row_data = [data_dict.get(header, "") for header in headers]
        sheet.append(row_data)

        try:
            workbook.save(EXCEL_FILE_PATH)
            return True # Sucesso
        except PermissionError:
            flash(f"Erro de Permissão ao salvar Excel. Feche o ficheiro '{EXCEL_FILE_PATH.name}' se estiver aberto.", "error")
            return False
        except Exception as save_error:
            flash(f"Erro ao salvar Excel: {save_error}", "error")
            return False
        finally:
            if workbook:
                workbook.close()

    except Exception as e:
        flash(f"Erro inesperado no processamento do Excel: {e}", "error")
        return False


def _print_pdf_logic(data_dict):
    """Lógica para gerar o PDF. Retorna o caminho do ficheiro ou None em caso de erro."""
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        font_family = "Arial"
        font_size = 11
        try:
            pdf.set_font(font_family, size=font_size)
        except RuntimeError:
             print("Aviso: Fonte Arial não encontrada, usando Helvetica.")
             font_family = "Helvetica"
             pdf.set_font(font_family, size=font_size)

        pdf.set_font(font_family, style='B', size=16)
        pdf.cell(0, 10, "Relatório de Ocorrência", ln=True, align='C')
        pdf.ln(10)
        pdf.set_font_size(font_size)

        # Iterar na ordem definida pelos headers para consistência
        headers = _get_excel_headers(data_dict)

        for key in headers:
            if key == 'Timestamp' or key == 'action': # Pular campos internos
                continue

            value = data_dict.get(key, "") # Obter valor do dicionário

            label = key.replace('_', ' ').replace('5w2h ', '5W2H - ').title()
            label = label.replace('Sgq', 'SGQ')

            pdf.set_font(font_family, style='B', size=font_size)
            label_width = pdf.get_string_width(f"{label}:") + 2
            pdf.cell(max(60, label_width), 7, f"{label}:", border=0)

            pdf.set_font(font_family, style='', size=font_size)

            # No formulário HTML, checkboxes vêm como 'true' (string) se marcados
            if key in ['atualizar_riscos', 'mudancas_sgq']:
                display_value = "Sim" if value == 'true' else "Não"
                pdf.cell(0, 7, display_value.encode('latin-1', 'replace').decode('latin-1'), ln=True, border=0)
            else:
                display_value = str(value)
                # Adicionar verificação para strings vazias ou None
                if display_value is None or display_value.strip() == "":
                    pdf.cell(0, 7, "-", ln=True, border=0) # Indicar campo vazio
                else:
                    pdf.multi_cell(0, 7, display_value.encode('latin-1', 'replace').decode('latin-1'), border=0, align='L')
                    pdf.ln(2) # Pequeno espaço após multi_cell para evitar sobreposição


        timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        tipo_ocorrencia_raw = data_dict.get("tipo", "Ocorrencia")
        tipo_ocorrencia_clean = "".join(c for c in tipo_ocorrencia_raw if c.isalnum() or c in (' ', '_', '-')).strip()
        tipo_ocorrencia_final = tipo_ocorrencia_clean.replace(' ', '_')[:30]
        if not tipo_ocorrencia_final: tipo_ocorrencia_final = "Ocorrencia"

        pdf_filename = f"{tipo_ocorrencia_final}_{timestamp_str}.pdf"
        pdf_filepath = PDF_SAVE_FOLDER / pdf_filename

        pdf.output(pdf_filepath)
        return pdf_filepath # Retorna o caminho completo do ficheiro gerado

    except Exception as e:
        flash(f"Erro inesperado ao gerar PDF: {e}", "error")
        import traceback
        print(f"Erro detalhado PDF:\n{traceback.format_exc()}")
        return None


# --- Rotas Flask ---

@app.route('/')
def index():
    """Exibe o formulário."""
    # Passa a lista de tipos para o template poder gerar o <select>
    return render_template('index.html', tipos_ocorrencia=TIPOS_OCORRENCIA)

@app.route('/submit', methods=['POST'])
def submit_report():
    """Processa os dados do formulário."""
    try:
        # Obter dados do formulário
        form_data = request.form.to_dict(flat=True)

        # Lógica para lidar com checkboxes (só vêm no form se marcados)
        form_data['atualizar_riscos'] = form_data.get('atualizar_riscos', 'false') # Vem como 'true' ou não vem
        form_data['mudancas_sgq'] = form_data.get('mudancas_sgq', 'false')       # Vem como 'true' ou não vem

        # Adicionar timestamp do servidor
        form_data['Timestamp'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Determinar qual botão foi clicado
        action = form_data.get('action') # Obtem 'save' ou 'pdf'

        if action == 'save':
            if _save_data_logic(form_data):
                flash("Dados salvos com sucesso no Excel!", "success")
            # Redireciona de volta para o formulário (mesmo se houve erro, o flash mostrará)
            return redirect(url_for('index'))

        elif action == 'pdf':
            pdf_path = _print_pdf_logic(form_data)
            if pdf_path and pdf_path.exists():
                try:
                    # Envia o ficheiro para download
                    return send_file(pdf_path, as_attachment=True)
                except Exception as send_err:
                    print(f"Erro ao enviar ficheiro PDF: {send_err}")
                    flash("Erro ao tentar enviar o PDF gerado.", "error")
                    # Mesmo se falhar o envio, redireciona para index
                    return redirect(url_for('index'))
            else:
                 # Erro na geração do PDF já foi "flashed" dentro de _print_pdf_logic
                 return redirect(url_for('index'))

        else:
            flash("Ação desconhecida.", "error")
            return redirect(url_for('index'))

    except Exception as e:
        flash(f"Erro inesperado no processamento do formulário: {e}", "error")
        import traceback
        print(f"Erro detalhado Form:\n{traceback.format_exc()}")
        return redirect(url_for('index'))

# --- Executar a Aplicação ---
if __name__ == '__main__':
    # host='0.0.0.0' torna acessível na rede local (opcional)
    # debug=True é útil para desenvolvimento, mas DESATIVE em produção
    app.run(debug=True, host='0.0.0.0')